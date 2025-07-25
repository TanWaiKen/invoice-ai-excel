import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from typing import List, Union, Dict
from datetime import datetime
import os

class ExcelCreator:
    def __init__(self):
        self.setup_styles()

    def setup_styles(self):
        """Setup Excel formatting styles"""
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.center_alignment = Alignment(horizontal="center", vertical="center")

    def create_excel_report(self, invoices: List[Union[Dict, object]], customers: List[object], output_path: str) -> str:
        """Create Excel report with only 2 sheets: Month Data + Price & Formula"""
        
        # Generate filename with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        month_name = self._get_dominant_month(invoices)
        
        if not output_path.endswith('.xlsx'):
            output_path = f"{output_path}_{month_name}_{timestamp}.xlsx"
        
        # Create workbook
        wb = openpyxl.Workbook()
        
        # Create only 2 sheets as requested
        self._create_main_data_sheet(wb, invoices, month_name)
        self._create_price_formula_sheet(wb, customers)
        
        # Create directory if it doesn't exist
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        
        # Save workbook
        wb.save(output_path)
        print(f"📁 Excel report saved: {output_path}")
        print(f"📊 Created 2 sheets: {month_name} (main data) + Price & Formula")
        
        return output_path

    def _create_main_data_sheet(self, workbook: openpyxl.Workbook, invoices: List[Union[Dict, object]], month_name: str):
        """Create main invoice data sheet with exact number of rows (no extra template rows)"""
        ws = workbook.active
        ws.title = f"{month_name}"
        
        # Enhanced headers matching your template structure
        headers = [
            "Date", "Invoices No", "Customer Name", "Weight (KG)", 
            "Prices Per Ton", "Total Bills", "Prices", "Company (RM)", 
            "Prices", "Worker (RM)"
        ]
        
        # Add headers with styling
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
            cell.font = Font(bold=True)
            cell.alignment = self.center_alignment
            cell.border = self.border
        
        # Sort invoices by invoice number
        try:
            sorted_invoices = sorted(invoices, key=lambda x: self._get_invoice_no_value(x))
        except:
            sorted_invoices = invoices
        
        print(f"🔍 Processing {len(sorted_invoices)} invoices for Excel sheet")
        
        # Add data with VLOOKUP formulas - ONLY ACTUAL DATA, NO DUPLICATES
        current_row = 2
        for i, invoice in enumerate(sorted_invoices):
            print(f"📝 Adding invoice {i+1}: {getattr(invoice, 'invoice_no', 'Unknown')} to row {current_row}")
            
            # Handle ONLY InvoiceData objects (no dict processing)
            date = getattr(invoice, "date", "")
            invoice_no = getattr(invoice, "invoice_no", "")
            customer_name = getattr(invoice, "customer_name", "")
            weight_kg = getattr(invoice, "weight_kg", 0)
            is_count = getattr(invoice, "is_count", True)
            
            # Special handling for Lain-lain
            if weight_kg != 1 and "lain" in customer_name.lower():
                print(f"🔧 Lain-lain detected for {customer_name}, setting weight to 1")
                weight_kg = 1
            
            # Determine formula based on is_count flag
            if is_count:
                # Count-based services: direct multiplication
                total_bills_formula = f"=D{current_row}*E{current_row}"
                company_formula = f"=D{current_row}*G{current_row}"
                worker_formula = f"=D{current_row}*I{current_row}"
            else:
                # Weight-based services: use /1000
                total_bills_formula = f"=D{current_row}*E{current_row}/1000"
                company_formula = f"=D{current_row}*G{current_row}/1000"
                worker_formula = f"=D{current_row}*I{current_row}/1000"
            
            row_data = [
                date,                                                                           # A: Date
                invoice_no,                                                                     # B: Invoice No  
                customer_name,                                                                  # C: Customer Name
                weight_kg,                                                                      # D: Weight (KG)
                f"=VLOOKUP(C{current_row},'Price & Formula'!A:E,2,FALSE)",                    # E: Prices Per Ton (VLOOKUP)
                total_bills_formula,                                                            # F: Total Bills (based on is_count)
                f"=VLOOKUP(C{current_row},'Price & Formula'!$A:$E,4,FALSE)",                  # G: Company Price (VLOOKUP)
                company_formula,                                                                # H: Company (RM) (based on is_count)
                f"=VLOOKUP(C{current_row},'Price & Formula'!$A:$E,5,FALSE)",                  # I: Worker Price (VLOOKUP)
                worker_formula                                                                  # J: Worker (RM) (based on is_count)
            ]
            
            # Add row data to worksheet
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=current_row, column=col_num)
                cell.value = value
                cell.border = self.border
                
                # Format columns
                if col_num == 4:  # Weight column - use general format, no decimals
                    if isinstance(value, (int, float)):
                        cell.number_format = '0'
                elif col_num >= 5:  # Other numeric columns
                    if isinstance(value, (int, float)):
                        cell.number_format = '#,##0.00'
        
            current_row += 1
        
        # Add grand total row (current_row is already the next empty row)
        total_row = current_row
        
        # "Grandtotal:" label
        label_cell = ws.cell(row=total_row, column=3, value="Grandtotal:")
        label_cell.font = Font(bold=True)
        label_cell.alignment = Alignment(horizontal="right")
        label_cell.border = self.border
        
        # Calculate range for totals (from row 2 to last data row)
        data_end_row = current_row - 1
        
        # Total Bills (Column F) - use SUM formula
        bills_total_cell = ws.cell(row=total_row, column=6, value=f"=SUM(F2:F{data_end_row})")
        bills_total_cell.font = Font(bold=True)
        bills_total_cell.number_format = '#,##0.00'
        bills_total_cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        bills_total_cell.border = self.border
        
        # Total Company (Column H) - use SUM formula
        company_total_cell = ws.cell(row=total_row, column=8, value=f"=SUM(H2:H{data_end_row})")
        company_total_cell.font = Font(bold=True)
        company_total_cell.number_format = '#,##0.00'
        company_total_cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        company_total_cell.border = self.border
        
        # Total Worker (Column J) - use SUM formula
        worker_total_cell = ws.cell(row=total_row, column=10, value=f"=SUM(J2:J{data_end_row})")
        worker_total_cell.font = Font(bold=True)
        worker_total_cell.number_format = '#,##0.00'
        worker_total_cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        worker_total_cell.border = self.border
        
        # Column width adjustment
        column_widths = {
            'A': 15, 'B': 12, 'C': 25, 'D': 12, 'E': 15,
            'F': 15, 'G': 15, 'H': 15, 'I': 15, 'J': 15
        }
        
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width
        
        print(f"✅ Main data sheet created: 1 header + {len(sorted_invoices)} data rows + 1 grand total row")
        print(f"📊 Total rows: {len(sorted_invoices) + 2}")  # header + data + grand total
        print(f"🔍 Final row structure: Header(1) + Data(2-{data_end_row}) + Total({total_row})")

    def _create_price_formula_sheet(self, workbook: openpyxl.Workbook, customers: List[object]):
        """Create Price & Formula sheet with customer data - showing actual formulas as TEXT not Excel formulas"""
        ws = workbook.create_sheet("Price & Formula")
        
        # Headers
        headers = ["Customer Name", "Prices Per Ton", "Formula", "Company", "Worker"]
        
        # Add headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = self.center_alignment
            cell.border = self.border
        
        # Add customer data
        for row_num, customer in enumerate(customers, 2):
            # Handle both dict and object types
            if hasattr(customer, 'name'):
                # Get actual formula text (not Excel formula!)
                formula_text = self._get_formula_display(customer.formula)
                
                row_data = [
                    customer.name,
                    customer.price_per_ton if customer.price_per_ton else 0,
                    formula_text,  # This should be TEXT, not Excel formula
                    customer.company_amount if customer.company_amount is not None else 0,
                    customer.worker_amount if customer.worker_amount is not None else 0
                ]
            else:
                formula_text = self._get_formula_display(customer.get("formula", ""))
                
                row_data = [
                    customer.get("name", ""),
                    customer.get("price_per_ton", 0),
                    formula_text,  # This should be TEXT, not Excel formula
                    customer.get("company_amount") if customer.get("company_amount") is not None else 0,
                    customer.get("worker_amount") if customer.get("worker_amount") is not None else 0
                ]
            
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num)
                
                # IMPORTANT: Force formula column to be TEXT, not Excel formula
                if col_num == 3 and isinstance(value, str) and value.startswith("="):
                    # Add apostrophe to force text format - this prevents Excel from treating it as formula
                    cell.value = "'" + value
                else:
                    cell.value = value
                
                cell.border = self.border
                
                # Format numbers
                if col_num >= 2 and col_num != 3 and isinstance(value, (int, float)):
                    cell.number_format = '#,##0.00'
        
        # Auto-adjust columns
        self._auto_adjust_columns(ws)
        
        print(f"✅ Price & Formula sheet created with {len(customers)} customers")
        print(f"📋 Formula column shows TEXT (not Excel formulas)")

    def _get_formula_display(self, formula: str) -> str:
        """Convert formula codes to display text - REMOVE equals sign to prevent Excel formula interpretation"""
        formula_mapping = {
            "WEIGHT_PRICE": "Weight x Price per ton/1000kg",  # NO equals sign!
            "Weight x Price per ton/1000kg": "Weight x Price per ton/1000kg",
            "ADDITIONAL": "Additional",
            "Additional": "Additional",
            "BAJA": "Baja",
            "Baja": "Baja",
            "JCB": "JCB",
            "LAIN_LAIN": "Lain Lain",
            "Lain Lain": "Lain Lain",
            "BAYARAN_GREDIR": "Lain Lain (Bayaran Gredir)",
            "Lain Lain (Bayaran Gredir)": "Lain Lain (Bayaran Gredir)",
            "MEMBAJA": "Membaja",
            "Membaja": "Membaja",
            "MEMOTONG_PELEPAH": "Memotong pelepah sawit",
            "Memotong pelepah sawit": "Memotong pelepah sawit",
            "MEMOTONG_PELEPAH_T": "Memotong pelepah sawit", 
            "Memotong pelepah sawit (T)": "Memotong pelepah sawit",  
            "MERACUN": "Meracun",
            "Meracun": "Meracun",
            "PAY_TO_GREDER": "Pay to Greder",
            "Pay to Greder": "Pay to Greder",
            "PENGANGKUTAN_LORI": "Pengangkutan Lori",
            "Pengangkutan Lori": "Pengangkutan Lori",
            "UPAH": "Upah",
            "Upah": "Upah"
        }
        
        # Remove any leading "=" if present and clean the formula
        clean_formula = formula.lstrip("=") if formula else ""
        
        # Get the display text (without equals sign)
        result = formula_mapping.get(formula, formula_mapping.get(clean_formula, clean_formula or "Weight x Price per ton/1000kg"))
        
        # If result starts with "=", remove it to prevent Excel formula interpretation
        if result.startswith("="):
            result = result[1:]
        
        return result

    def _get_invoice_no_value(self, invoice) -> int:
        """Extract numeric part from invoice number for sorting"""
        try:
            if isinstance(invoice, dict):
                invoice_no = str(invoice.get("invoice_no", ""))
            else:
                invoice_no = str(getattr(invoice, "invoice_no", ""))
            
            # Extract numbers from invoice string
            import re
            numbers = re.findall(r'\d+', invoice_no)
            return int(numbers[0]) if numbers else 0
        except:
            return 0

    def _get_dominant_month(self, invoices: List[Union[Dict, object]]) -> str:
        """Get the most common month from invoices"""
        if not invoices:
            return datetime.now().strftime("%B")
        
        try:
            # Extract months from dates
            months = []
            for invoice in invoices:
                try:
                    if isinstance(invoice, dict):
                        date_str = invoice.get("date", "")
                    else:
                        date_str = getattr(invoice, "date", "")
                        
                    if date_str:
                        date_obj = datetime.strptime(date_str, "%d/%m/%Y")
                        months.append(date_obj.strftime("%B"))
                except:
                    continue
            
            if months:
                # Get most common month
                from collections import Counter
                return Counter(months).most_common(1)[0][0]
            else:
                return datetime.now().strftime("%B")
        except:
            return datetime.now().strftime("%B")

    def _auto_adjust_columns(self, worksheet):
        """Auto-adjust column widths"""
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width