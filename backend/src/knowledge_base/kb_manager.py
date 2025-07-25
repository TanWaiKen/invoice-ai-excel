from multiprocessing import process
import os
import re
import sys
import unicodedata
import openpyxl
from typing import List, Optional, Dict, Any
from langchain_community.vectorstores import Chroma
from langchain.schema import Document
from fuzzywuzzy import fuzz, process

# Only need embeddings for vector search
try:
    from langchain_ollama import OllamaEmbeddings
    OLLAMA_AVAILABLE = True
except ImportError:
    try:
        from langchain_community.embeddings import OllamaEmbeddings
        OLLAMA_AVAILABLE = True
    except ImportError:
        OLLAMA_AVAILABLE = False

from src.models.data_models import Customer

class KnowledgeBaseManager:
    def __init__(self):
        """Initialize Knowledge Base Manager with Vector Database for RAG"""
        self.customers: List[Customer] = []
        self.vector_store = None
        self.retriever = None
        
        # Vector database configuration
        self.vector_db_path = "../knowledge_base/data/vector_db"
        os.makedirs(self.vector_db_path, exist_ok=True)
        
        # Initialize embeddings for vector search only
        self.embeddings = self._initialize_embeddings()
    
    def _initialize_embeddings(self):
        """Initialize Ollama embedding model for vector search"""
        if not OLLAMA_AVAILABLE:
            return None
        
        try:
            embeddings = OllamaEmbeddings(
                model="nomic-embed-text",
                base_url="http://localhost:11434"
            )
            # Test the connection
            test_embed = embeddings.embed_query("test")
            return embeddings
        except Exception:
            return None

    def load_from_excel(self, excel_path: str):
        """Load customer data from Excel and create vector database"""
        if not os.path.exists(excel_path):
            raise FileNotFoundError(f"Excel file not found: {excel_path}")
        
        try:
            wb = openpyxl.load_workbook(excel_path, read_only=False)
            
            # Find the right worksheet
            ws = None
            for sheet_name in wb.sheetnames:
                if any(keyword in sheet_name.lower() for keyword in ['price', 'formula']):
                    ws = wb[sheet_name]
                    break
            
            if not ws:
                wb.close()
                raise ValueError(f"No suitable worksheet found in {excel_path}")
            
            # Find headers
            header_row = None
            col_mappings = {}
            
            for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=1), 1):
                for col_idx, cell in enumerate(row):
                    if cell.value and isinstance(cell.value, str):
                        header_text = cell.value.lower().strip()
                        
                        if any(keyword in header_text for keyword in ['customer', 'name']):
                            col_mappings['name'] = col_idx
                            header_row = row_idx
                        elif any(keyword in header_text for keyword in ['price', 'ton']):
                            col_mappings['price_per_ton'] = col_idx
                            header_row = row_idx
                        elif 'formula' in header_text:
                            col_mappings['formula'] = col_idx
                            header_row = row_idx
                        elif 'company' in header_text and 'worker' not in header_text:
                            col_mappings['company_amount'] = col_idx
                            header_row = row_idx
                        elif 'worker' in header_text:
                            col_mappings['worker_amount'] = col_idx
                            header_row = row_idx
                
                if header_row:
                    break
            
            if not header_row:
                raise ValueError("Could not find header row in Excel file")
            
            # Clear existing customers
            self.customers = []
            
            # Read customer data
            for row in ws.iter_rows(min_row=header_row + 1):
                try:
                    # Get name
                    if 'name' not in col_mappings:
                        continue
                        
                    name_cell = row[col_mappings['name']]
                    if not name_cell.value or str(name_cell.value).strip() == "":
                        continue
                    
                    name = str(name_cell.value).strip()
                    
                    # Get price_per_ton
                    price_per_ton = 0
                    if 'price_per_ton' in col_mappings:
                        price_cell = row[col_mappings['price_per_ton']]
                        if price_cell.value is not None:
                            try:
                                price_per_ton = float(price_cell.value)
                            except:
                                price_per_ton = 0
                    
                    # Get formula
                    formula = "WEIGHT_PRICE"
                    if 'formula' in col_mappings:
                        formula_cell = row[col_mappings['formula']]
                        if formula_cell.value:
                            formula = str(formula_cell.value).strip()
                    
                    # Get company_amount
                    company_amount = None
                    if 'company_amount' in col_mappings:
                        company_cell = row[col_mappings['company_amount']]
                        if company_cell.value is not None:
                            try:
                                company_value = float(company_cell.value)
                                if company_value > 0:
                                    company_amount = company_value
                            except:
                                company_amount = None
                    
                    # Get worker_amount
                    worker_amount = None
                    if 'worker_amount' in col_mappings:
                        worker_cell = row[col_mappings['worker_amount']]
                        
                        if worker_cell.value is not None:
                            try:
                                worker_value = float(worker_cell.value)
                                if worker_value > 0:
                                    worker_amount = round(worker_value, 2)
                            except:
                                if price_per_ton > 0 and company_amount is not None:
                                    worker_amount = round(price_per_ton - company_amount, 2)
                    
                    # Final calculation if still None
                    if worker_amount is None and price_per_ton > 0 and company_amount is not None:
                        worker_amount = round(price_per_ton - company_amount, 2)
                    
                    # Create customer
                    customer = Customer(
                        name=name,
                        price_per_ton=price_per_ton,
                        formula=formula,
                        company_amount=company_amount,
                        worker_amount=worker_amount
                    )
                    
                    self.customers.append(customer)
                    
                except Exception:
                    continue
            
            wb.close()
            
            if len(self.customers) == 0:
                raise ValueError("No valid customer data found in Excel file")
            
            # Create vector database
            self.create_vector_database()
            
        except Exception as e:
            raise Exception(f"Error loading Excel: {e}")

    def create_vector_database(self):
        """Create vector database for RAG retrieval"""
        try:
            if not self.embeddings or not self.customers:
                return
            
            # Clean up existing
            self.cleanup_vector_store()
            
            # Create new path to avoid conflicts
            import uuid
            import time
            timestamp = int(time.time())
            self.vector_db_path = f"../knowledge_base/data/vector_db_{timestamp}_{uuid.uuid4().hex[:8]}"
            os.makedirs(self.vector_db_path, exist_ok=True)
            
            # Prepare documents for vector search
            documents = []
            for customer in self.customers:
                content = f"Customer: {customer.name} | Price per ton: RM{customer.price_per_ton} | Formula: {customer.formula}"
                
                if customer.company_amount:
                    content += f" | Company amount: RM{customer.company_amount}"
                
                if customer.worker_amount:
                    content += f" | Worker amount: RM{customer.worker_amount}"
                
                metadata = {
                    "name": customer.name,
                    "price_per_ton": str(customer.price_per_ton),
                    "formula": customer.formula,
                    "company_amount": str(customer.company_amount) if customer.company_amount is not None else "0",
                    "worker_amount": str(customer.worker_amount) if customer.worker_amount is not None else "0"
                }
                
                documents.append(Document(page_content=content, metadata=metadata))
            
            # Create vector store
            self.vector_store = Chroma.from_documents(
                documents=documents,
                embedding=self.embeddings,
                persist_directory=self.vector_db_path,
                collection_name="customers"
            )
            
            self.retriever = self.vector_store.as_retriever(search_kwargs={"k": 10})
            
        except Exception:
            self.vector_store = None
            self.retriever = None

    def cleanup_vector_store(self):
        """Clean up existing vector store"""
        try:
            if hasattr(self, 'vector_store') and self.vector_store:
                if hasattr(self.vector_store, '_client'):
                    try:
                        del self.vector_store._client
                    except:
                        pass
                
                del self.vector_store
                self.vector_store = None
            
            import gc
            gc.collect()
            
        except Exception:
            pass

    def find_customer_matches(self, extracted_name: str, k: int = 5) -> List[Dict]:
        """Find top k customer matches using fuzzy matching for LLM processing"""
        try:
            if not self.customers:
                return []
            
            # Clean the extracted name
            cleaned_extracted = self.clean_name_for_matching(extracted_name)
            
            # Get all customer names and their cleaned versions
            customer_data = []
            for customer in self.customers:
                cleaned_name = self.clean_name_for_matching(customer.name)
                customer_data.append({
                    "original_name": customer.name,
                    "cleaned_name": cleaned_name,
                    "customer": customer
                })
            
            # Perform fuzzy matching
            cleaned_names = [item["cleaned_name"] for item in customer_data]
            matches = process.extract(cleaned_extracted, cleaned_names, scorer=fuzz.ratio, limit=k)
            
            # Build result list with customer details
            results = []
            for match_name, score in matches:
                # Find the corresponding customer
                for item in customer_data:
                    if item["cleaned_name"] == match_name:
                        customer = item["customer"]
                        results.append({
                            "customer_name": customer.name,
                            "price_per_ton": customer.price_per_ton,
                            "formula": customer.formula,
                            "company_amount": customer.company_amount,
                            "worker_amount": customer.worker_amount,
                            "confidence_score": score / 100.0,
                            "metadata": {
                                "name": customer.name,
                                "price_per_ton": str(customer.price_per_ton),
                                "formula": customer.formula,
                                "company_amount": str(customer.company_amount) if customer.company_amount is not None else "0",
                                "worker_amount": str(customer.worker_amount) if customer.worker_amount is not None else "0"
                            }
                        })
                        break
            
            return results
            
        except Exception:
            return []

    def vector_search_customers(self, query: str, k: int = 5) -> List[Dict]:
        """Search customers using vector similarity"""
        try:
            if not self.vector_store:
                return []
            
            docs = self.vector_store.similarity_search_with_score(query, k=k)
            
            results = []
            for doc, score in docs:
                result = {
                    "content": doc.page_content,
                    "metadata": doc.metadata,
                    "similarity_score": 1.0 - score,
                    "customer_name": doc.metadata.get("name", "Unknown")
                }
                results.append(result)
            
            return results
            
        except Exception:
            return []
    
    def clean_name_for_matching(self, name: str) -> str:
        """Clean name for better matching - remove titles and company suffixes"""
        if not name:
            return ""
        
        cleaned = name.lower().strip()
        
        # Remove personal titles AND company suffixes
        prefixes_to_remove = ['mr.', 'mrs.', 'ms.', 'en.', 'pn.', 'dr.', 'prof.']
        suffixes_to_remove = ['sdn bhd', 'sdn.bhd.', 'bhd', 'pte ltd', 'ltd', 'inc', 'corp']
        
        # Remove prefixes
        for prefix in prefixes_to_remove:
            if cleaned.startswith(prefix):
                cleaned = cleaned[len(prefix):].strip()
        
        # Remove suffixes
        for suffix in suffixes_to_remove:
            if cleaned.endswith(suffix):
                cleaned = cleaned[:-len(suffix)].strip()
        
        # Remove special characters but keep spaces and important punctuation
        cleaned = re.sub(r'[^\w\s\u4e00-\u9fff@&().-]', ' ', cleaned)
        
        # Normalize unicode
        cleaned = unicodedata.normalize('NFKC', cleaned)
        
        # Remove extra spaces
        cleaned = ' '.join(cleaned.split())
        
        return cleaned
    
    def get_all_customers(self) -> List[Customer]:
        """Get all loaded customers"""
        return self.customers
    
    def get_vector_store_info(self) -> Dict[str, Any]:
        """Get information about the vector store"""
        try:
            if not self.vector_store:
                return {"status": "No vector store created"}
            
            return {
                "status": "Vector store active",
                "document_count": len(self.customers),
                "embedding_model": "nomic-embed-text",
                "vector_db_path": self.vector_db_path,
                "rag_ready": self.retriever is not None
            }
            
        except Exception as e:
            return {"status": f"Error getting vector store info: {e}"}
    

    def __del__(self):
        """Destructor to clean up resources"""
        try:
            self.cleanup_vector_store()
        except Exception:
            pass
    
    def search_similar_customers(self, query: str, top_k: int = 10) -> List[Dict]:
        """Enhanced search using fuzzy matching with complete customer data including price"""
        try:
            if not self.customers:
                return []
            
            # Use find_customer_matches which already returns price_per_ton
            matches = self.find_customer_matches(query, k=top_k)
            
            # Ensure all required fields are present
            for match in matches:
                if 'price_per_ton' not in match:
                    match['price_per_ton'] = 0.0
                if 'confidence_score' not in match:
                    match['confidence_score'] = 0.0
                    
            print(f"🔍 Vector search results with prices:")
            for match in matches:
                print(f"   📊 {match['customer_name']} | Price: RM{match['price_per_ton']} | Confidence: {match['confidence_score']:.3f}")
            
            return matches
            
        except Exception as e:
            print(f"❌ Search error: {e}")
            return []

    def get_customer_by_name(self, customer_name: str):
        """Get customer by exact name match"""
        for customer in self.customers:
            if customer.name == customer_name:
                return customer
        return None

    def add_customer_variant(self, customer_name: str, price_per_ton: float, formula: str = "WEIGHT_PRICE") -> bool:
        """Add new customer variant to the knowledge base"""
        try:
            from src.models.data_models import Customer
            
            # Check if customer already exists
            existing = self.get_customer_by_name(customer_name)
            if existing:
                print(f"⚠️ Customer {customer_name} already exists")
                return False
            
            # Create new customer
            new_customer = Customer(
                name=customer_name,
                price_per_ton=price_per_ton,
                formula=formula,
                company_amount=None,
                worker_amount=None
            )
            
            # Add to list
            self.customers.append(new_customer)
            
            # Recreate vector database
            self.create_vector_database()
            
            print(f"✅ Added new customer: {customer_name} | RM{price_per_ton}")
            return True
            
        except Exception as e:
            print(f"❌ Failed to add customer: {e}")
            return False
